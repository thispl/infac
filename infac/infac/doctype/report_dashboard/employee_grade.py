from __future__ import unicode_literals
from os import name, sched_getaffinity
from time import time
import frappe
from frappe.monitor import start
from frappe.utils import cstr, add_days, date_diff, getdate, format_date
from frappe import _, append_hook, bold
from frappe.utils.csvutils import UnicodeWriter, read_csv_content
from frappe.utils.data import format_date
from frappe.utils.file_manager import get_file
from frappe.model.document import Document
from frappe.utils.background_jobs import enqueue

from datetime import date, timedelta, datetime
import openpyxl
from openpyxl import Workbook

import openpyxl
import xlrd
import re
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import GradientFill, PatternFill
from six import BytesIO, string_types
#Employee Grade report 
@frappe.whitelist()
def download():
    filename = 'Employee Grade'
    test = build_xlsx_response(filename)


def make_xlsx(data, sheet_name=None, wb=None, column_widths=None):
    args = frappe.local.form_dict
    column_widths = column_widths or []
    if wb is None:
        wb = openpyxl.Workbook()
    ws = wb.create_sheet(sheet_name, 0)
    #column header_1
    ws.append(['EMPLOYEES ATTENDANCE '])
    #column header_2
    header_1 = ['','','','']
    dates = get_dates(args)
    header_2= ['S.NO','EMP.CODE','EMP NAME','DEPARTMEN']
    rows = ['IN','OUT','STATUS','NIGHT','OT','LATE']
    for d in dates:
        day = datetime.strptime(str(d),'%Y-%m-%d').strftime('%d-%m')
        header_1.extend(['','','','','',day])
        header_2.extend(rows)
    header_2.extend(['Present','Absent','Leave','OT','Permission','Late','Night','Total L and A'])
    ws.append(header_1)
    ws.append(header_2)
    employee = frappe.db.get_all('Employee',{'status':'Active','employee_category':('in',['Master Worker','Workers Grade 2'])},['*'])
    i = 1
    for emp in employee:
        emp_id = emp.employee
        name = emp.employee_name
        dept = emp.department
        row = [i,emp_id,name,dept]
        status = ''
        n_sft = ''
        total_ot = 0
        total_present = 0
        total_absent = 0
        total_leave = 0
        total_night = 0
        permission = 0
        for dt in dates:
            att = frappe.db.get_value('Attendance',{'employee':emp_id,'attendance_date':dt},['in_time','out_time','status','ot_hrs','late_entry','shift'])
            pr = frappe.db.get_value('Permission Request',{'employee_id':emp_id,'permission_date':dt},['session'])
            if pr:
                if pr == 'First Half':
                    permission = permission + 0.5 
                elif pr == 'Second Half':
                    permission = permission + 0.5 
            if att:
                if att[5] == 'C':
                    n_sft = 1
                    total_night = total_night + n_sft
                if att[2]:
                    if att[2] == 'Present':
                        status = 'P'
                        total_present = total_present + 1
                    elif att[2] == 'Absent':
                        status = 'A'
                        total_absent = total_absent + 1
                    elif att[2] == 'On Leave':
                        status = 'L'
                        total_leave = total_leave + 1
                row.extend([att[0],att[1],status,n_sft,att[3],att[4]])
                total_ot =total_ot + att[3]
            else:
                row.extend(['','','','','',''])
        la = total_absent + total_leave + permission
        row.extend([total_present,total_absent,total_leave,total_ot,permission,'',total_night,la])
        ws.append(row)
        i = i+1
    #     date = get_dates(args)
    #     for d in date:
    #         day = datetime.strptime(str(d),'%Y-%m-%d').strftime('%d-%m')
    #         att = frappe.db.get_value('Attendance',{'employee':emp.name,'attendance_date':d},['in_time','out_time','status'])
    #         if att:
    #             if att.status == 'Present':
    #                 status = "P"
    #             else:
    #                 status = "A"
    #             row.append(att.in_time)
    #             row.append(att.out_time)          
    #             row.append(att.status)
    #             header.append(row)
    # ws.append(header)  
        # for date in dates:
        #     attendance = frappe.db.get_value('Attendance',{'attendance':emp.name,'attendance_date':date},['in_time','out_time','status'])
        #     if attendance:
        #         if attendance.status == 'Present':
        #             status = 'P'
        #         else:
        #             status = 'A'
        #         row.append(attendance.in_time)
        #         row.append(attendance.out_time)
        # ws.append(row)       
                    
    # department = frappe.db.get_all('Department')
    # for dept in department:
    #     employee = frappe.get_all('Employee',{'department':dept.name},['*'])
    #     for emp in employee:
    #         row = [emp.name,emp.employee_name,emp.department]
    #         for date in dates:datre
    #             att = frappe.db.get_value('Attendance',{'attendance':emp.name,'attendance_date':date},['in_time','out_time','status'],as_dict=True)
    #             print(att)
               


    xlsx_file = BytesIO()
    wb.save(xlsx_file)
    return xlsx_file


def build_xlsx_response(filename):
    xlsx_file = make_xlsx(filename)
    frappe.response['filename'] = filename + '.xlsx'
    frappe.response['filecontent'] = xlsx_file.getvalue()
    frappe.response['type'] = 'binary'	

def get_dates(args):
    no_of_days = date_diff(add_days(args.to_date, 1), args.from_date)
    dates = [add_days(args.from_date, i) for i in range(0, no_of_days)]
    return dates


